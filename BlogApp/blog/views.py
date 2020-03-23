from django.conf.global_settings import MEDIA_ROOT
from django.shortcuts import render
from .models import BIP,BWS,Oekonomie
from openpyxl import load_workbook
from django.utils import timezone
from django import forms
from django.contrib.auth.models import User
from .form import SelectForm, ImportForm



from django.http import HttpResponseRedirect
# Create your views here.
from django.conf import settings
import pdb
from django.db.models.fields.files import ImageFieldFile, FileField


def handle_uploaded_file(f):
    with open('media/images/name.jpg', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

# def post_list(request):
#     #posts = Post.objects.filter(publishedDate__lte=timezone.now()).order_by('publishedDate')
#     #sString = request.__getattribute__('val')
#     form = SelectForm(request.POST,request.FILES)
#     if form.is_valid():
#         search = form.cleaned_data['firstname']
#         posts = Post.objects.filter(title__contains=search)
#         posts = posts.order_by('title')
#         uploadedFile = request.FILES['image']
#         handle_uploaded_file(uploadedFile)
#
#
#
#     return render(request, 'blog/post_list.html', {'posts':posts})

def index(request):
    if request.method == 'POST':
        pdb.set_trace()
        # create a form instance and populate it with data from the request:
        form = SelectForm(request.POST)
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            s = form.cleaned_data['firstname']

            return HttpResponseRedirect('/plist/')
    # if a GET (or any other method) we'll create a blank form
    else:
        form = SelectForm()
    return render(request,'blog/index.html',{'form':form})

def diagram(request):
    return render(request, 'blog/diagram.html')
def importData(request):
    if request.method == 'POST':

        # create a form instance and populate it with data from the request:
        form = ImportForm(request.POST,request.FILES)
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:

            if request.user.is_authenticated and request.user.username == 'cxd':
                #pdb.set_trace()
                s = form.cleaned_data['importFile']
                wb = load_workbook(s)
                sheet = wb['Lausitzdatenbank_mn_cp']
                rIdx = 5
                yearCol = 1
                row = next(sheet.iter_rows(min_row=rIdx, max_row=rIdx, min_col=1, max_col=7, values_only=True))
                while row[0] is not None:

                    BIPObj= BIP.objects.create(Einwohner=row[2],Erwerbstaetiger=row[3])
                    BWSObj = BWS.objects.create(Einwohner=row[4], Erwerbstaetiger=row[5],ErwerbstaetigerLandForstFischerei=row[6])
                    OekObj =  Oekonomie.objects.create(Jahr = row[0],Landkreis = row[1], BIPField = BIPObj,BWSField = BWSObj)
                    rIdx = rIdx + 1
                    row = next(sheet.iter_rows(min_row=rIdx, max_row=rIdx, min_col=1, max_col=7, values_only=True))

                wb.save('file.xlsx')
                return HttpResponseRedirect('/importOK')
    # if a GET (or any other method) we'll create a blank form
    else:
        form = ImportForm()
    return render(request, 'blog/import.html',{'form':form})
def importOK(request):
    return render(request, 'blog/importOK.html')